from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QListWidget, QListWidgetItem,
    QGroupBox, QFormLayout, QDateEdit, QCheckBox, QComboBox, QSpinBox, QTableWidget,
    QTableWidgetItem, QMessageBox, QFileDialog, QScrollArea, QFrame, QHeaderView,
    QMenu, QInputDialog
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor, QPalette
from student_system.core.database import Database
import json
from datetime import datetime, timedelta, time, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from collections import defaultdict


class ExamScheduler(QWidget):
    def __init__(self, user, parent=None):
        super().__init__(parent)
        self.user = user
        # İstisna süreleri (ders_id -> dakika) saklamak için
        self.lesson_exceptions = {}

        self._init_ui()
        self._apply_modern_styles()
        self._load_lessons()

    def _init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)

        self._create_header(main_layout)

        content_layout = QHBoxLayout()
        content_layout.setSpacing(20)

        self._create_left_panel(content_layout)
        self._create_right_panel(content_layout)

        main_layout.addLayout(content_layout)

    def _create_header(self, parent_layout):
        title = QLabel(f"{self.user['bolum_adi']} - Sınav Programı")
        title.setAlignment(Qt.AlignLeft)
        title.setStyleSheet("""
            font-size: 28px;
            font-weight: 600;
            color: #2c3e50;
            margin-bottom: 5px;
        """)
        parent_layout.addWidget(title)

    def _create_left_panel(self, parent_layout):
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)

        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setSpacing(20)

        self._create_lesson_selection_group(left_layout)
        self._create_date_settings_group(left_layout)
        self._create_exam_settings_group(left_layout)
        self._create_action_buttons(left_layout)

        left_layout.addStretch()
        scroll_area.setWidget(left_widget)
        parent_layout.addWidget(scroll_area, 2)

    def _create_lesson_selection_group(self, parent_layout):
        group = QGroupBox("Ders Seçimi (İstisna Süre için Sağ Tıkla)")  # <-- YENİ
        group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 2px solid #e0e0e0;
                border-radius: 10px;
                padding: 20px 15px 15px 15px;
                font-weight: 600;
                font-size: 15px;
                color: #2c3e50;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
            }
        """)
        layout = QVBoxLayout(group)
        layout.setSpacing(12)

        self.lesson_list = QListWidget()
        self.lesson_list.setMinimumHeight(250)
        self.lesson_list.setSelectionMode(QListWidget.NoSelection)
        self.lesson_list.setStyleSheet("""
            QListWidget {
                background-color: #fafbfc;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                padding: 8px;
                font-size: 14px;
            }
            QListWidget::item {
                padding: 10px;
                border-radius: 6px;
                margin: 2px 0;
            }
            QListWidget::item:hover {
                background-color: #e8f0fe;
            }
        """)

        # --- YENİ: Sağ Tık Menüsü ---
        self.lesson_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.lesson_list.customContextMenuRequested.connect(self._show_lesson_context_menu)
        # --- BİTTİ ---

        layout.addWidget(self.lesson_list)

        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        self.btn_all = QPushButton("Tümünü Seç")
        self.btn_all.clicked.connect(lambda: self._toggle_all(True))
        self.btn_all.setStyleSheet("""
            QPushButton {
                background-color: #667eea;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #5568d3;
            }
            QPushButton:pressed {
                background-color: #4c5cba;
            }
        """)

        self.btn_none = QPushButton("Tümünü Kaldır")
        self.btn_none.clicked.connect(lambda: self._toggle_all(False))
        self.btn_none.setStyleSheet("""
            QPushButton {
                background-color: #e0e0e0;
                color: #2c3e50;
                border: none;
                padding: 10px 20px;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #cbd5e0;
            }
            QPushButton:pressed {
                background-color: #a0aec0;
            }
        """)

        button_layout.addWidget(self.btn_all)
        button_layout.addWidget(self.btn_none)
        layout.addLayout(button_layout)

        parent_layout.addWidget(group)

    # --- YENİ FONKSİYON: Sağ Tık Menüsünü Göster ---
    def _show_lesson_context_menu(self, pos):
        item = self.lesson_list.itemAt(pos)
        if not item:
            return

        menu = QMenu(self)

        set_action = menu.addAction("İstisna Sınav Süresi Belirle...")
        reset_action = menu.addAction("Varsayılan Süreye Dön")

        set_action.triggered.connect(lambda: self._set_exception_duration(item))
        reset_action.triggered.connect(lambda: self._reset_exception_duration(item))

        menu.exec_(self.lesson_list.mapToGlobal(pos))

    # --- YENİ FONKSİYON: İstisna Süresi Ayarla ---
    def _set_exception_duration(self, item):
        ders_data = item.data(Qt.UserRole)
        ders_id = ders_data['ders_id']
        ders_adi = ders_data['ders_adi']

        current_duration = self.lesson_exceptions.get(ders_id, self.spin_duration.value())

        duration, ok = QInputDialog.getInt(
            self,
            "İstisna Süre",
            f"'{ders_adi}' için sınav süresi (dk):",
            value=current_duration,
            min=15,
            max=240,
            step=15
        )

        if ok:
            self.lesson_exceptions[ders_id] = duration
            # Orijinal metni koruyarak istisnayı ekle
            original_text = f"{ders_data['ders_kodu']} – {ders_data['ders_adi']} (Sınıf: {ders_data['sinif']})"
            item.setText(f"{original_text}  [İSTİSNA: {duration} dk]")
            item.setForeground(QColor("#d97706"))  # Rengi değiştir
            font = item.font()
            font.setBold(True)
            item.setFont(font)

    # --- YENİ FONKSİYON: İstisna Süreyi Sıfırla ---
    def _reset_exception_duration(self, item):
        ders_data = item.data(Qt.UserRole)
        ders_id = ders_data['ders_id']

        if ders_id in self.lesson_exceptions:
            del self.lesson_exceptions[ders_id]

        original_text = f"{ders_data['ders_kodu']} – {ders_data['ders_adi']} (Sınıf: {ders_data['sinif']})"
        item.setText(original_text)
        item.setForeground(QColor("#2c3e50"))  # Rengi sıfırla
        font = item.font()
        font.setBold(False)
        item.setFont(font)

    def _create_date_settings_group(self, parent_layout):
        group = QGroupBox("Tarih ve Gün Ayarları")
        group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 2px solid #e0e0e0;
                border-radius: 10px;
                padding: 20px 15px 15px 15px;
                font-weight: 600;
                font-size: 15px;
                color: #2c3e50;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
            }
        """)
        form_layout = QFormLayout(group)
        form_layout.setSpacing(15)

        self.date_from = QDateEdit(QDate.currentDate())
        self.date_from.setCalendarPopup(True)
        self.date_from.setStyleSheet("""
            QDateEdit {
                padding: 10px;
                border: 2px solid #e0e0e0;
                border-radius: 6px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:focus {
                border: 2px solid #667eea;
            }
        """)

        self.date_to = QDateEdit(QDate.currentDate().addDays(14))
        self.date_to.setCalendarPopup(True)
        self.date_to.setStyleSheet("""
            QDateEdit {
                padding: 10px;
                border: 2px solid #e0e0e0;
                border-radius: 6px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:focus {
                border: 2px solid #667eea;
            }
        """)

        form_layout.addRow("Başlangıç:", self.date_from)
        form_layout.addRow("Bitiş:", self.date_to)

        days_label = QLabel("Sınav Günleri:")
        days_label.setStyleSheet("font-weight: 500; color: #4a5568; font-size: 14px;")
        form_layout.addRow(days_label)

        days_layout = QHBoxLayout()
        days_layout.setSpacing(8)

        self.chk_mon = self._create_day_checkbox("Pzt", True)
        self.chk_tue = self._create_day_checkbox("Sal", True)
        self.chk_wed = self._create_day_checkbox("Çar", True)
        self.chk_thu = self._create_day_checkbox("Per", True)
        self.chk_fri = self._create_day_checkbox("Cum", True)
        self.chk_sat = self._create_day_checkbox("Cmt", False)
        self.chk_sun = self._create_day_checkbox("Paz", False)

        for checkbox in [self.chk_mon, self.chk_tue, self.chk_wed, self.chk_thu,
                         self.chk_fri, self.chk_sat, self.chk_sun]:
            days_layout.addWidget(checkbox)

        form_layout.addRow(days_layout)
        parent_layout.addWidget(group)

    def _create_day_checkbox(self, text, checked):
        checkbox = QCheckBox(text)
        checkbox.setChecked(checked)
        checkbox.setStyleSheet("""
            QCheckBox {
                background-color: white;
                border: 2px solid #e0e0e0;
                border-radius: 6px;
                padding: 8px 10px;
                font-size: 13px;
                font-weight: 500;
            }
            QCheckBox:hover {
                border: 2px solid #667eea;
            }
            QCheckBox:checked {
                background-color: #667eea;
                color: white;
                border: 2px solid #667eea;
            }
            QCheckBox::indicator {
                width: 0px;
                height: 0px;
            }
        """)
        return checkbox

    def _create_exam_settings_group(self, parent_layout):
        group = QGroupBox("Sınav Ayarları")
        group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 2px solid #e0e0e0;
                border-radius: 10px;
                padding: 20px 15px 15px 15px;
                font-weight: 600;
                font-size: 15px;
                color: #2c3e50;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
            }
        """)
        form_layout = QFormLayout(group)
        form_layout.setSpacing(15)

        self.cmb_type = QComboBox()
        self.cmb_type.addItems(["Vize", "Final", "Bütünleme"])
        self.cmb_type.setStyleSheet("""
            QComboBox {
                padding: 10px;
                border: 2px solid #e0e0e0;
                border-radius: 6px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #667eea;
            }
            QComboBox::drop-down {
                border: none;
                padding-right: 10px;
            }
        """)

        self.spin_duration = QSpinBox()
        self.spin_duration.setRange(30, 240)
        self.spin_duration.setValue(75)
        self.spin_duration.setSuffix(" dakika")
        self.spin_duration.setStyleSheet("""
            QSpinBox {
                padding: 10px;
                border: 2px solid #e0e0e0;
                border-radius: 6px;
                font-size: 14px;
                background-color: white;
            }
            QSpinBox:focus {
                border: 2px solid #667eea;
            }
        """)

        self.spin_break = QSpinBox()
        self.spin_break.setRange(0, 120)
        self.spin_break.setValue(15)
        self.spin_break.setSuffix(" dakika")
        self.spin_break.setStyleSheet("""
            QSpinBox {
                padding: 10px;
                border: 2px solid #e0e0e0;
                border-radius: 6px;
                font-size: 14px;
                background-color: white;
            }
            QSpinBox:focus {
                border: 2px solid #667eea;
            }
        """)

        self.chk_no_overlap = QCheckBox("Sınavlar çakışmasın")
        self.chk_no_overlap.setStyleSheet("""
            QCheckBox {
                padding: 8px;
                font-weight: 500;
                font-size: 14px;
            }
        """)
        self.chk_no_overlap.setToolTip(
            "İşaretlenirse, aynı anda sadece bir sınav yapılır.\n"
            "İşaretlenmezse, öğrenci çakışması olmayan sınavlar aynı saate atanır."
        )

        form_layout.addRow("Sınav Türü:", self.cmb_type)
        form_layout.addRow("Varsayılan Süre:", self.spin_duration)
        form_layout.addRow("Sınav Arası:", self.spin_break)
        form_layout.addRow(self.chk_no_overlap)

        parent_layout.addWidget(group)

    def _create_action_buttons(self, parent_layout):
        self.btn_generate = QPushButton("Programı Oluştur")
        self.btn_generate.setMinimumHeight(50)
        self.btn_generate.clicked.connect(self._on_generate_clicked)
        self.btn_generate.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                padding: 12px;
                border-radius: 8px;
                font-size: 15px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #059669;
            }
            QPushButton:pressed {
                background-color: #047857;
            }
        """)

        self.export_button = QPushButton("Excel Olarak İndir")
        self.export_button.setMinimumHeight(45)
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #f59e0b;
                color: white;
                border: none;
                padding: 12px;
                border-radius: 8px;
                font-size: 14px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #d97706;
            }
            QPushButton:pressed {
                background-color: #b45309;
            }
        """)

        parent_layout.addWidget(self.btn_generate)
        parent_layout.addWidget(self.export_button)

    def _create_right_panel(self, parent_layout):
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setSpacing(15)

        preview_label = QLabel("Program Önizlemesi")
        preview_label.setStyleSheet("""
            font-size: 18px;
            font-weight: 600;
            color: #2c3e50;
            padding: 5px 0;
        """)
        right_layout.addWidget(preview_label)

        self.preview = QTableWidget(0, 6)
        self.preview.setHorizontalHeaderLabels([
            "Ders Kodu", "Ders Adı", "Tarih", "Saat", "Sınıf", "Derslik"
        ])
        self.preview.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.preview.setSelectionBehavior(QTableWidget.SelectRows)
        self.preview.setSelectionMode(QTableWidget.SingleSelection)
        self.preview.setAlternatingRowColors(True)
        self.preview.setShowGrid(False)
        self.preview.verticalHeader().setVisible(False)
        self.preview.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 2px solid #e0e0e0;
                border-radius: 10px;
                font-size: 14px;
            }
            QTableWidget::item {
                padding: 12px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #667eea;
                color: white;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 12px;
                border: none;
                border-bottom: 2px solid #e0e0e0;
                font-weight: 600;
                color: #2c3e50;
                font-size: 13px;
            }
            QTableWidget::item:alternate {
                background-color: #fafbfc;
            }
        """)

        right_layout.addWidget(self.preview)
        parent_layout.addWidget(right_widget, 3)

    def _apply_modern_styles(self):
        self.setStyleSheet("""
            QWidget {
                background-color: #f5f7fa;
                color: #2c3e50;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QLabel {
                color: #4a5568;
                font-size: 14px;
                font-weight: 500;
            }
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                background: #f5f7fa;
                width: 10px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical {
                background: #cbd5e0;
                border-radius: 5px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background: #a0aec0;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
        """)

    def _load_lessons(self):
        # Programı her yeniden yüklediğimizde istisna listesini temizle
        self.lesson_exceptions.clear()

        rows = Database.execute_query("""
            SELECT ders_id, ders_kodu, ders_adi, sinif
            FROM dersler
            WHERE bolum_id = %s AND aktif = true
            ORDER BY sinif, ders_kodu
        """, (self.user["bolum_id"],))

        self.lesson_list.clear()
        for r in rows or []:
            item = QListWidgetItem(f"{r['ders_kodu']} – {r['ders_adi']} (Sınıf: {r['sinif']})")
            item.setData(Qt.UserRole, r)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked)
            self.lesson_list.addItem(item)

    def _toggle_all(self, checked):
        state = Qt.Checked if checked else Qt.Unchecked
        for i in range(self.lesson_list.count()):
            self.lesson_list.item(i).setCheckState(state)

    def _collect_constraints(self):
        selected = []
        for i in range(self.lesson_list.count()):
            item = self.lesson_list.item(i)
            if item.checkState() == Qt.Checked:
                ders_data = item.data(Qt.UserRole).copy()  # Kopyasını al
                # İstisna süreyi dersin verisine ekle
                ders_data['istisna_sure'] = self.lesson_exceptions.get(ders_data['ders_id'])
                selected.append(ders_data)

        if not selected:
            raise ValueError("Programa dahil edilecek en az bir ders seçmelisiniz.")

        # Döküman kuralı: Aynı sınıfları bir arada tutmak için sırala
        selected.sort(key=lambda d: (d['sinif'], d['ders_kodu']))

        d_from = self.date_from.date().toPyDate()
        d_to = self.date_to.date().toPyDate()

        if d_to < d_from:
            raise ValueError("Bitiş tarihi başlangıçtan önce olamaz.")

        allowed_weekdays = []
        checkboxes = [self.chk_mon, self.chk_tue, self.chk_wed, self.chk_thu,
                      self.chk_fri, self.chk_sat, self.chk_sun]

        for idx, checkbox in enumerate(checkboxes):
            if checkbox.isChecked():
                allowed_weekdays.append(idx)

        if not allowed_weekdays:
            raise ValueError("En az bir gün seçmelisiniz.")

        return {
            "bolum_id": self.user["bolum_id"],
            "sinav_turu": self.cmb_type.currentText(),
            "varsayilan_sure": self.spin_duration.value(),
            "varsayilan_bekleme": self.spin_break.value(),
            "baslangic_tarihi": d_from,
            "bitis_tarihi": d_to,
            "allowed_weekdays": allowed_weekdays,
            "ayni_anda_sinav_engelle": self.chk_no_overlap.isChecked(),
            "secili_dersler": selected
        }

    def _on_generate_clicked(self):
        try:
            constraints = self._collect_constraints()

            schedule = self._perform_scheduling(constraints)

            self.assign_exam_rooms()
            self._update_preview(schedule)

            QMessageBox.information(
                self,
                "Başarılı",
                f"{len(schedule)} adet sınav başarıyla planlandı!"
            )
        except ValueError as e:
            QMessageBox.warning(self, "Uyarı", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Program oluşturulurken hata: {str(e)}")

    def _add_time(self, current_time, minutes_to_add):
        """ Bir datetime.time nesnesine dakika ekler, tarihi yok sayar. """
        dummy_date = datetime(2000, 1, 1, current_time.hour, current_time.minute, current_time.second)
        new_datetime = dummy_date + timedelta(minutes=minutes_to_add)
        return new_datetime.time()

    # --- YARDIMCI FONKSİYON 2 (DEĞİŞİKLİK YOK, AMA GEREKLİ) ---
    def _check_student_conflict(self, ders_ogrencileri, sinav_tarihi, sinav_baslangic, sinav_bitis,
                                ogrenci_programi):
        """ Verilen dersin öğrencilerinin, o zaman aralığında başka bir sınavı var mı? """

        for ogrenci_id in ders_ogrencileri:
            for kayitli_tarih, kayitli_baslangic, kayitli_bitis in ogrenci_programi[ogrenci_id]:

                if sinav_tarihi != kayitli_tarih:
                    continue  # Farklı gün, çakışma olamaz

                # Zaman aralığı çakışma kontrolü
                # (StartA < EndB) and (EndA > StartB)
                if (sinav_baslangic < kayitli_bitis) and (sinav_bitis > kayitli_baslangic):
                    return True  # Çakışma var!

        return False  # Çakışma yok

    # --- YARDIMCI FONKSİYON 3 (DEĞİŞİKLİK YOK, AMA GEREKLİ) ---
    def _update_student_schedule(self, ders_ogrencileri, sinav_tarihi, sinav_baslangic, sinav_bitis_with_break,
                                 ogrenci_programi):
        """ Öğrencilerin programına bu sınavı (ve bekleme süresini) ekler. """

        kayit = (sinav_tarihi, sinav_baslangic, sinav_bitis_with_break)
        for ogrenci_id in ders_ogrencileri:
            ogrenci_programi[ogrenci_id].append(kayit)

        # --- GÜNCELLENMİŞ ÇEKİRDEK FONKSİYON ---
        # Not: _add_time, _check_student_conflict, ve _update_student_schedule
        #      yardımcı fonksiyonlarının class'ınızda tanımlı olduğunu varsayıyorum.

    def _perform_scheduling(self, c):
        conn = Database.get_connection()
        cur = conn.cursor()

        # 1. Eski sınav kayıtlarını temizle
        cur.execute("""
            DELETE FROM Sinavlar 
            WHERE bolum_id = %s AND sinav_turu = %s
        """, (c["bolum_id"], c["sinav_turu"]))

        # 2. Gerekli verileri hazırla

        ders_ogrenci_map = defaultdict(set)
        ders_idler = [d['ders_id'] for d in c['secili_dersler']]
        if ders_idler:
            placeholders = ','.join(['%s'] * len(ders_idler))
            cur.execute(f"""
                SELECT ders_id, ogrenci_id 
                FROM ogrencidersleri 
                WHERE ders_id IN ({placeholders})
            """, tuple(ders_idler))
            for row in cur.fetchall():
                # Ders ID ve Öğrenci ID'nin int olduğundan emin ol
                ders_ogrenci_map[int(row[0])].add(int(row[1]))

        # 3. Zamanlama için veri yapılarını kur

        ogrenci_programi = defaultdict(list)
        gun_timeline = {}

        current_date = c["baslangic_tarihi"]
        SINAV_BASLANGIC_SAATI = time(9, 0)
        SINAV_BITIS_SAATI = time(17, 0)

        while current_date <= c["bitis_tarihi"]:
            if current_date.weekday() in c["allowed_weekdays"]:
                gun_timeline[current_date] = SINAV_BASLANGIC_SAATI
            current_date += timedelta(days=1)

        if not gun_timeline:
            raise ValueError("Seçilen tarih aralığında sınav yapılabilecek gün bulunamadı.")

        gun_sinif_sayisi = defaultdict(lambda: defaultdict(int))
        slot_programi = defaultdict(list)
        schedule = []

        # 4. Dersleri Planlama Döngüsü
        for ders in c["secili_dersler"]:
            ders_suresi = ders['istisna_sure'] or c['varsayilan_sure']
            ders_ogrencileri = ders_ogrenci_map.get(ders['ders_id'], set())
            placed = False

            sorted_gunler = sorted(
                gun_timeline.keys(),
                key=lambda tarih: gun_sinif_sayisi[tarih][ders['sinif']]
            )

            for tarih in sorted_gunler:

                # --- A. ADIM: MEVCUT SLOTLARA YERLEŞMEYİ DENE (GREEDY) ---
                if not c["ayni_anda_sinav_engelle"]:
                    gunun_slotlari = [
                        (saat, dersler) for (t, saat), dersler in slot_programi.items() if t == tarih
                    ]

                    for saat, slot_dersleri in gunun_slotlari:
                        en_uzun_sure = max(
                            (d['istisna_sure'] or c['varsayilan_sure']) for d in slot_dersleri
                        )
                        if ders_suresi > en_uzun_sure:
                            continue

                        bitis_saati = self._add_time(saat, en_uzun_sure)
                        bitis_saati_ara_dahil = self._add_time(bitis_saati, c['varsayilan_bekleme'])

                        conflict = self._check_student_conflict(
                            ders_ogrencileri, tarih, saat, bitis_saati_ara_dahil, ogrenci_programi
                        )

                        if not conflict:
                            schedule.append((ders, tarih, saat))
                            slot_programi[(tarih, saat)].append(ders)
                            self._update_student_schedule(
                                ders_ogrencileri, tarih, saat, bitis_saati_ara_dahil, ogrenci_programi
                            )
                            gun_sinif_sayisi[tarih][ders['sinif']] += 1
                            placed = True
                            break  # Slot arama (iç) döngüsünden çık

                if placed:
                    break  # Gün arama (dış) döngüsünden de çık

                # --- B. ADIM: YENİ SLOT OLUŞTUR (GÜNÜN SONUNA EKLE) ---

                # 'Günde 1 Sınav' kuralı kontrolü
                if gun_sinif_sayisi[tarih][ders['sinif']] > 0:
                    continue  # Bu güne bu sınıftan zaten 1 sınav atandı, başka güne bak

                baslangic_saati = gun_timeline[tarih]

                bitis_saati = self._add_time(baslangic_saati, ders_suresi)
                bitis_saati_ara_dahil = self._add_time(bitis_saati, c['varsayilan_bekleme'])

                if bitis_saati > SINAV_BITIS_SAATI:
                    continue

                conflict = self._check_student_conflict(
                    ders_ogrencileri, tarih, baslangic_saati, bitis_saati_ara_dahil, ogrenci_programi
                )

                if not conflict:
                    schedule.append((ders, tarih, baslangic_saati))
                    slot_programi[(tarih, baslangic_saati)].append(ders)
                    self._update_student_schedule(
                        ders_ogrencileri, tarih, baslangic_saati, bitis_saati_ara_dahil, ogrenci_programi
                    )
                    gun_timeline[tarih] = bitis_saati_ara_dahil
                    gun_sinif_sayisi[tarih][ders['sinif']] += 1
                    placed = True
                    break  # Gün arama (dış) döngüsünden çık

            # *** BAŞARISIZLIK DURUMU KONTROLÜ (GENEL HALE GETİRİLDİ) ***
            if not placed:
                cur.close()
                conn.close()

                # Başarısız olan dersin bilgilerini al
                failing_sinif = ders['sinif']
                failing_ders_adi = ders['ders_adi']
                gun_sayisi = len(gun_timeline)

                # Başarısız olan dersin sınıfına ait planlanacak toplam ders sayısını bul
                failing_sinif_ders_sayisi = sum(1 for d in c['secili_dersler'] if d['sinif'] == failing_sinif)

                # Bu sınıftaki tüm öğrencilerin tüm dersleri alıp almadığını kontrol et
                # (Bu, 1. Sınıf senaryosu gibi özel durumlar için)
                is_full_overlap = True  # Varsayılan olarak tam çakışma varsay
                ogrenciler_bu_sinifta = set()

                for d in c['secili_dersler']:
                    if d['sinif'] == failing_sinif:
                        ogrenciler_bu_sinifta.update(ders_ogrenci_map.get(d['ders_id'], set()))

                if ogrenciler_bu_sinifta:
                    for d in c['secili_dersler']:
                        if d['sinif'] == failing_sinif:
                            # Eğer bu sınıftaki öğrencilerin tamamı bu dersi almıyorsa,
                            # tam çakışma yoktur.
                            if not ogrenciler_bu_sinifta.issubset(ders_ogrenci_map.get(d['ders_id'], set())):
                                is_full_overlap = False
                                # break # Not: Bu kontrol %100 doğru olmayabilir ama iyi bir tahmin

                # Eğer "günde 1 sınav" kuralı yüzünden sığmıyorsa:
                if (failing_sinif_ders_sayisi > gun_sayisi) and is_full_overlap:
                    raise ValueError(
                        f"GÜN YETERSİZ! '{failing_ders_adi}' dersi atanamadı.\n\n"
                        f"Sebep: {failing_sinif}. sınıf için {failing_sinif_ders_sayisi} adet çakışan sınav planlanıyor "
                        f"ancak 'günde bir sınav' kuralı için sadece {gun_sayisi} gün mevcut.\n\n"
                        "Lütfen tarih aralığını genişletin."
                    )
                else:
                    # Diğer tüm çakışmalar (alttan ders, vb.) veya zaman darlığı için
                    raise ValueError(
                        f"'{failing_ders_adi}' dersi için uygun zaman bulunamadı! \n\n"
                        f"Sebep: Tarih aralığı, {failing_sinif}. sınıfın diğer dersleri ve "
                        "diğer sınıflarla olan öğrenci çakışmalarını (örn. alttan ders alma) "
                        "karşılamak için yetersiz.\n\n"
                        "Lütfen tarih aralığını genişletin."
                    )

        # 5. Planlanan sınavları veritabanına yaz
        for ders, tarih, saat_obj in schedule:
            cur.execute("""
                INSERT INTO Sinavlar 
                (ders_id, bolum_id, sinav_turu, sinav_tarihi, sinav_saati, durum)
                VALUES (%s, %s, %s, %s, %s, 'Planlandı')
            """, (ders["ders_id"], c["bolum_id"], c["sinav_turu"],
                  tarih, saat_obj))

        conn.commit()
        cur.close()
        conn.close()

        return schedule

    def _update_preview(self, schedule):
        self.preview.setRowCount(0)

        conn = Database.get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT 
                s.sinav_id,
                d.ders_kodu,
                d.ders_adi,
                d.sinif,
                s.sinav_tarihi,
                s.sinav_saati,
                COALESCE(STRING_AGG(dl.derslik_adi, ', '), 'Atanamadı') AS derslikler
            FROM Sinavlar s
            JOIN Dersler d ON s.ders_id = d.ders_id
            LEFT JOIN SinavDerslikleri sd ON s.sinav_id = sd.sinav_id
            LEFT JOIN Derslikler dl ON sd.derslik_id = dl.derslik_id
            WHERE s.bolum_id = %s AND s.sinav_turu = %s
            GROUP BY s.sinav_id, d.ders_kodu, d.ders_adi, d.sinif, s.sinav_tarihi, s.sinav_saati
            ORDER BY s.sinav_tarihi, s.sinav_saati
        """, (self.user["bolum_id"], self.cmb_type.currentText()))

        rows = cur.fetchall()

        for i, row in enumerate(rows):
            sinav_id, kod, adi, sinif, tarih, saat, derslikler = row

            self.preview.insertRow(i)
            self.preview.setItem(i, 0, QTableWidgetItem(kod))
            self.preview.setItem(i, 1, QTableWidgetItem(adi))
            self.preview.setItem(i, 2, QTableWidgetItem(tarih.strftime("%d.%m.%Y")))
            self.preview.setItem(i, 3, QTableWidgetItem(saat.strftime("%H:%M")))
            self.preview.setItem(i, 4, QTableWidgetItem(str(sinif)))
            self.preview.setItem(i, 5, QTableWidgetItem(derslikler))

            self.preview.setRowHeight(i, 50)

        cur.close()
        conn.close()

    def export_to_excel(self):
        try:
            path, _ = QFileDialog.getSaveFileName(
                self, "Excel Dosyası Kaydet", "",
                "Excel Files (*.xlsx)"
            )
            if not path:
                return

            conn = Database.get_connection()
            cur = conn.cursor()

            sinav_turu = self.cmb_type.currentText()

            cur.execute("""
                SELECT 
                    s.sinav_tarihi AS tarih,
                    s.sinav_saati AS saat,
                    d.ders_adi AS ders_adi,
                    COALESCE(ou.ad_soyad, 'Atanmamış') AS ogretim_elemani,
                    COALESCE(STRING_AGG(dl.derslik_adi, ', '), 'Atanmamış') AS derslik
                FROM Sinavlar s
                JOIN Dersler d ON s.ders_id = d.ders_id
                LEFT JOIN OgretimUyeleri ou ON d.hoca_id = ou.hoca_id
                LEFT JOIN SinavDerslikleri sd ON s.sinav_id = sd.sinav_id
                LEFT JOIN Derslikler dl ON sd.derslik_id = dl.derslik_id
                WHERE s.bolum_id = %s AND s.sinav_turu = %s
                GROUP BY s.sinav_id, d.ders_adi, ou.ad_soyad, s.sinav_tarihi, s.sinav_saati
                ORDER BY s.sinav_tarihi ASC, s.sinav_saati ASC
            """, (self.user["bolum_id"], sinav_turu))

            rows = cur.fetchall()

            if not rows:
                QMessageBox.warning(self, "Uyarı", "Oluşturulmuş sınav programı yok!")
                return

            df = pd.DataFrame(rows, columns=["Tarih", "Saat", "Ders Adı", "Öğretim Elemanı", "Derslik"])

            wb = Workbook()
            ws = wb.active
            ws.title = f"{sinav_turu} Programı"

            turuncu_baslik = "F4B084"
            turuncu_gun = "FCE4D6"
            beyaz = "FFFFFF"

            border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            ws.merge_cells("A1:E1")
            ws["A1"] = f"{self.user['bolum_adi']} {sinav_turu.upper()} SINAV PROGRAMI"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].fill = PatternFill("solid", fgColor=turuncu_baslik)

            headers = ["Tarih", "Sınav Saati", "Ders Adı", "Öğretim Elemanı", "Derslik"]
            ws.append(headers)

            for c in range(1, 6):
                cell = ws.cell(row=2, column=c)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=turuncu_baslik)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            renk_sira = 0
            satir = 3

            for tarih, grup in df.groupby("Tarih"):
                gun_rengi = PatternFill("solid", fgColor=(turuncu_gun if renk_sira % 2 == 0 else beyaz))
                renk_sira += 1

                first_row = satir

                for _, row in grup.iterrows():
                    ws.cell(satir, 2, value=row["Saat"].strftime("%H:%M"))
                    ws.cell(satir, 3, row["Ders Adı"])
                    ws.cell(satir, 4, row["Öğretim Elemanı"])
                    ws.cell(satir, 5, row["Derslik"])

                    for c in range(2, 6):
                        cell = ws.cell(row=satir, column=c)
                        cell.fill = gun_rengi
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border

                    satir += 1

                ws.merge_cells(start_row=first_row, start_column=1, end_row=satir - 1, end_column=1)
                date_cell = ws.cell(first_row, 1)
                date_cell.value = tarih.strftime("%d.%m.%Y")
                date_cell.font = Font(bold=True)
                date_cell.border = border
                date_cell.fill = gun_rengi
                date_cell.alignment = Alignment(text_rotation=90, horizontal="center", vertical="center")

            widths = [15, 15, 40, 30, 25]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            wb.save(path)
            QMessageBox.information(self, "Başarılı", f"Excel dosyası kaydedildi:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Hata oluştu:\n{str(e)}")

    def assign_exam_rooms(self):
        try:
            conn = Database.get_connection()
            cur = conn.cursor()

            cur.execute("""
                SELECT derslik_id, derslik_adi, kapasite
                FROM Derslikler
                WHERE aktif = true AND bolum_id = %s
                ORDER BY kapasite DESC
            """, (self.user["bolum_id"],))
            derslikler = cur.fetchall()

            if not derslikler:
                QMessageBox.warning(self, "Uyarı", "Aktif derslik bulunamadı!")
                cur.close()
                conn.close()
                return

            sinav_turu = self.cmb_type.currentText()
            cur.execute("""
                SELECT s.sinav_id, s.ders_id, s.sinav_tarihi, s.sinav_saati, d.ders_adi
                FROM Sinavlar s
                JOIN Dersler d ON s.ders_id = d.ders_id
                WHERE s.bolum_id = %s AND s.sinav_turu = %s
                ORDER BY s.sinav_tarihi, s.sinav_saati
            """, (self.user["bolum_id"], sinav_turu))
            sinavlar = cur.fetchall()

            if not sinavlar:
                # Bu artık bir hata değil, program daha oluşmamış olabilir
                # QMessageBox.warning(self, "Uyarı", "Henüz sınav bulunamadı!")
                cur.close()
                conn.close()
                return

            # --- YENİ: Önceki derslik atamalarını temizle ---
            sinav_idler = [s[0] for s in sinavlar]
            if sinav_idler:
                cur.execute(f"""
                    DELETE FROM SinavDerslikleri 
                    WHERE sinav_id IN ({','.join(['%s'] * len(sinav_idler))})
                """, tuple(sinav_idler))
            # --- BİTTİ ---

            cur.execute("""
                SELECT ders_id, COUNT(ogrenci_id) as ogrenci_sayisi
                FROM OgrenciDersleri
                GROUP BY ders_id
            """)
            ogrenci_sayilari = {r[0]: r[1] for r in cur.fetchall()}

            # derslik_programi: derslik_id -> {tarih -> [saat, ...]}
            derslik_programi = defaultdict(lambda: defaultdict(list))

            toplam_atama = 0
            uyari_listesi = []

            for sinav_id, ders_id, tarih, saat, ders_adi in sinavlar:
                # Tip dönüşümlerini garantile (veritabanından farklı gelebilir)
                saat_time = saat if isinstance(saat, time) else datetime.strptime(str(saat), "%H:%M:%S").time()
                tarih_date = tarih if isinstance(tarih, type(date.today())) else tarih.date()

                ogrenci_sayisi = ogrenci_sayilari.get(ders_id, 0)

                if ogrenci_sayisi == 0:
                    uyari_listesi.append(f"⚠️ {ders_adi}: Öğrenci bulunamadı, atlanıyor.")
                    continue

                gereken_derslikler = self._select_optimal_classrooms(
                    derslikler=derslikler,
                    ogrenci_sayisi=ogrenci_sayisi,
                    tarih_date=tarih_date,
                    saat_time=saat_time,
                    derslik_programi=derslik_programi,
                )

                if not gereken_derslikler:
                    uyari_listesi.append(
                        f"❌ {ders_adi} ({ogrenci_sayisi} öğrenci): "
                        f"Uygun derslik bulunamadı! Tüm derslikler dolu veya kapasite yetersiz."
                    )
                    continue

                atanan_derslik_adlari = []
                toplam_atanan_kapasite = 0

                for derslik_id, derslik_adi, kapasite in gereken_derslikler:
                    # Bu dersliği bu (tarih, saat) için meşgul olarak işaretle
                    derslik_programi[derslik_id][tarih_date].append(saat_time)

                    cur.execute("""
                        INSERT INTO SinavDerslikleri (sinav_id, derslik_id)
                        VALUES (%s, %s)
                        ON CONFLICT DO NOTHING
                    """, (sinav_id, derslik_id))

                    atanan_derslik_adlari.append(f"{derslik_adi}({kapasite})")
                    toplam_atanan_kapasite += kapasite
                    toplam_atama += 1

                derslik_str = " + ".join(atanan_derslik_adlari)

                if toplam_atanan_kapasite < ogrenci_sayisi:
                    uyari_listesi.append(
                        f"⚠️ {ders_adi}: {ogrenci_sayisi} öğrenci, "
                        f"atanan kapasite: {toplam_atanan_kapasite} → {derslik_str} (KAPASİTE YETERSİZ!)"
                    )
                else:
                    # Başarılı atama (çok kalabalık yapmaması için loglama kapatılabilir)
                    # uyari_listesi.append(f"✅ {ders_adi} ({ogrenci_sayisi} ögr) → {derslik_str}")
                    pass

            conn.commit()
            cur.close()
            conn.close()

            mesaj = f"Toplam {toplam_atama} derslik ataması yapıldı."

            if uyari_listesi:
                # Sadece ❌ ve ⚠️ içeren uyarıları göster
                onemli_uyarilar = [u for u in uyari_listesi if "❌" in u or "⚠️" in u]
                if onemli_uyarilar:
                    mesaj += f"\n\n⚠️ {len(onemli_uyarilar)} ÖNEMLİ UYARI:\n" + "\n".join(onemli_uyarilar[:10])
                    if len(onemli_uyarilar) > 10:
                        mesaj += f"\n... ve {len(onemli_uyarilar) - 10} uyarı daha"

            # Çok fazla uyarı penceresi çıkmaması için bu kapatılabilir
            # QMessageBox.information(self, "Derslik Atama Tamamlandı", mesaj)

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Derslik atama hatası: {str(e)}")

    def _select_optimal_classrooms(self, derslikler, ogrenci_sayisi, tarih_date,
                                   saat_time, derslik_programi):

        # O tarih ve saatte boş olan derslikleri bul
        uygun_derslikler = []
        for derslik_id, derslik_adi, kapasite in derslikler:
            gunluk_program = derslik_programi[derslik_id].get(tarih_date, [])

            if saat_time in gunluk_program:
                continue  # Bu derslik bu saatte dolu

            uygun_derslikler.append((derslik_id, derslik_adi, kapasite))

        if not uygun_derslikler:
            return []

        # Döküman kuralı: "Derslik kullanımı en aza indirgeyecek şekilde yerleştirme"
        # Bu, öğrenci sayısını karşılayana kadar en büyük derslikleri seçerek sağlanır.

        secilen_derslikler = []
        kalan_ogrenci = ogrenci_sayisi

        # Uygun derslikler zaten kapasiteye göre (büyükten küçüğe) sıralı
        for derslik in uygun_derslikler:
            derslik_id, derslik_adi, kapasite = derslik
            secilen_derslikler.append(derslik)
            kalan_ogrenci -= kapasite

            if kalan_ogrenci <= 0:
                break  # Yeterli kapasiteye ulaşıldı

        return secilen_derslikler